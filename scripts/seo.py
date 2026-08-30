#!/usr/bin/env python3
"""
seo — разбор выгрузок из Search Console, Метрики, Вебмастера и GA4.

Сырые выгрузки — это тысячи строк. Скрипт читает их с диска и печатает
сводку строк на пятьдесят. В контекст попадают выводы, а не таблицы.

  1. Кладёшь файлы в seo/inbox/ (csv, tsv, xlsx, zip — как выгрузилось)
  2. python3 scripts/seo.py digest
  3. Claude Code читает сводку и пишет бриф

  python3 scripts/seo.py files      что лежит в inbox и как распозналось
  python3 scripts/seo.py digest     полная сводка
  python3 scripts/seo.py queries    только запросы
  python3 scripts/seo.py pages      только страницы + сверка с репозиторием
  python3 scripts/seo.py archive    убрать разобранное в seo/archive/<дата>/
"""
import csv, io, json, os, re, shutil, signal, subprocess, sys, zipfile
from datetime import date
from pathlib import Path

try:
    signal.signal(signal.SIGPIPE, signal.SIG_DFL)
except (AttributeError, ValueError):
    pass


def root() -> Path:
    try:
        return Path(subprocess.check_output(["git", "rev-parse", "--show-toplevel"],
                                            stderr=subprocess.DEVNULL, text=True).strip())
    except Exception:
        return Path.cwd()


ROOT = root()
INBOX = ROOT / "seo" / "inbox"
ARCHIVE = ROOT / "seo" / "archive"
SITE = "mir-doma.pro"

# Синонимы колонок: русский и английский интерфейсы, разные сервисы.
COLS = {
    "query":  ["запрос", "query", "поисковый запрос", "ключевое слово", "фраза",
               "поисковая фраза"],
    "page":   ["страница", "page", "url", "адрес страницы", "landing page",
               "целевая страница", "страница входа"],
    "clicks": ["клики", "clicks", "переходы", "визиты", "сеансы", "sessions",
               "пользователи", "users"],
    "impr":   ["показы", "impressions", "показов"],
    "ctr":    ["ctr", "кликабельность"],
    "pos":    ["позиция", "position", "средняя позиция", "average position"],
    "date":   ["дата", "date", "день"],
    "depth":  ["глубина просмотра", "страниц за визит", "pages per session",
               "просмотров на визит"],
    "bounce": ["отказы", "bounce rate", "процент отказов"],
}


def norm(s: str) -> str:
    return re.sub(r"[\s_\-]+", " ", (s or "").strip().lower()).strip()


def match_col(header):
    """header -> {ключ: индекс}. Точное совпадение важнее подстроки."""
    out = {}
    hn = [norm(h) for h in header]
    for key, names in COLS.items():
        for i, h in enumerate(hn):
            if h in names:
                out[key] = i
                break
        if key not in out:
            for i, h in enumerate(hn):
                if any(n in h for n in names):
                    out[key] = i
                    break
    return out


def num(v):
    if v is None:
        return 0.0
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace("%", "")
    s = s.replace(",", ".")
    if not s or s in ("-", "—"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


# ---------------------------------------------------------------- чтение


def read_rows(path: Path):
    """Возвращает список строк-списков. Понимает csv/tsv/xlsx/zip."""
    suf = path.suffix.lower()
    if suf == ".zip":
        out = []
        with zipfile.ZipFile(path) as z:
            for name in z.namelist():
                if name.lower().endswith((".csv", ".tsv")):
                    raw = z.read(name).decode("utf-8-sig", "replace")
                    out.append((name, list(csv.reader(io.StringIO(raw),
                                                      delimiter=sniff(raw)))))
        return out
    if suf == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        out = []
        for ws in wb.worksheets:
            rows = [[c if c is not None else "" for c in r]
                    for r in ws.iter_rows(values_only=True)]
            out.append((ws.title, rows))
        return out
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    return [(path.name, list(csv.reader(io.StringIO(raw), delimiter=sniff(raw))))]


def sniff(raw: str) -> str:
    head = raw[:4000]
    return max([",", ";", "\t"], key=lambda d: head.count(d))


def find_header(rows):
    """GA4 и Метрика кладут мусор перед шапкой — ищем первую распознаваемую строку."""
    for i, r in enumerate(rows[:25]):
        if len(r) < 2:
            continue
        cols = match_col([str(c) for c in r])
        if ("query" in cols or "page" in cols) and \
           ("clicks" in cols or "impr" in cols):
            return i, cols
    for i, r in enumerate(rows[:25]):
        cols = match_col([str(c) for c in r])
        if len(cols) >= 2:
            return i, cols
    return None, {}


def guess_kind(name: str, cols: dict) -> str:
    n = name.lower()
    if "impr" in cols and "pos" in cols:
        base = "GSC"
    elif "impr" in cols:
        base = "Вебмастер"
    else:
        base = "Метрика/GA4"
    seg = ""
    if any(w in n for w in ("image", "картин", "изображ")):
        seg = " (картинки)"
    elif any(w in n for w in ("web", "веб")):
        seg = " (веб)"
    if "query" in cols:
        return f"{base}: запросы{seg}"
    if "page" in cols:
        return f"{base}: страницы{seg}"
    return f"{base}: прочее{seg}"


def load_inbox():
    """[(файл, вид, cols, строки-словари)]"""
    if not INBOX.exists():
        INBOX.mkdir(parents=True, exist_ok=True)
        sys.exit(f"Создал {INBOX.relative_to(ROOT)} — положи туда выгрузки и повтори.")
    files = [p for p in sorted(INBOX.iterdir())
             if p.suffix.lower() in (".csv", ".tsv", ".xlsx", ".zip")]
    if not files:
        sys.exit(f"В {INBOX.relative_to(ROOT)} пусто. Положи выгрузки (csv/xlsx/zip).")
    out = []
    for p in files:
        try:
            sheets = read_rows(p)
        except Exception as e:
            print(f"# не прочитал {p.name}: {e}", file=sys.stderr)
            continue
        for sheet_name, rows in sheets:
            hi, cols = find_header(rows)
            if hi is None or not cols:
                continue
            data = []
            for r in rows[hi + 1:]:
                if not any(str(c).strip() for c in r):
                    continue
                d = {}
                for k, i in cols.items():
                    d[k] = r[i] if i < len(r) else ""
                data.append(d)
            if data:
                label = f"{p.name}" + (f" / {sheet_name}" if sheet_name != p.name else "")
                out.append((label, guess_kind(f"{p.name} {sheet_name}", cols), cols, data))
    return out


# ---------------------------------------------------------------- анализ


def repo_slugs():
    idx = ROOT / ".mdi-index.json"
    if not idx.exists():
        return {}
    try:
        return json.loads(idx.read_text(encoding="utf-8"))["docs"]
    except Exception:
        return {}


def slug_of(url: str) -> str:
    m = re.search(r"//[^/]*/([a-z0-9\-]+)/?", str(url))
    return m.group(1) if m else ""


def cmd_files(_=None):
    for label, kind, cols, data in load_inbox():
        print(f"{len(data):>6} строк  {kind:<28} {label}")
        print(f"        колонки: {', '.join(sorted(cols))}")


def _queries(data):
    out = []
    for d in data:
        q = str(d.get("query", "")).strip()
        if not q:
            continue
        out.append((q, num(d.get("clicks")), num(d.get("impr")), num(d.get("pos"))))
    return out


def cmd_queries(_=None):
    blocks = [(l, k, d) for l, k, c, d in load_inbox() if "query" in c]
    if not blocks:
        print("# выгрузок с запросами не нашёл")
        return
    for label, kind, data in blocks:
        qs = _queries(data)
        if not qs:
            continue
        clicks = sum(q[1] for q in qs)
        impr = sum(q[2] for q in qs)
        ctr = clicks / impr * 100 if impr else 0
        print(f"\n=== {kind} — {label}")
        print(f"запросов {len(qs)}, кликов {clicks:.0f}, показов {impr:.0f}, CTR {ctr:.2f}%")

        # Топ-3 без кликов: интент, который забирают AI-выжимка и карусель картинок.
        dead = [q for q in qs if q[3] and q[3] <= 3 and q[1] == 0 and q[2] >= 10]
        if dead:
            dead.sort(key=lambda x: -x[2])
            print(f"\nТоп-3 и ноль кликов ({len(dead)}) — показы есть, трафика нет:")
            for q, c, i, p in dead[:8]:
                print(f"  поз {p:>4.1f}  показов {i:>6.0f}  {q}")

        # Позиции 5-20 с показами: дотянуть дешевле, чем писать новое.
        near = [q for q in qs if q[3] and 5 <= q[3] <= 20 and q[2] >= 20]
        if near:
            near.sort(key=lambda x: -x[2])
            print(f"\nБлизко к топу, 5-20 позиция ({len(near)}) — усилить существующее:")
            for q, c, i, p in near[:10]:
                cr = c / i * 100 if i else 0
                print(f"  поз {p:>4.1f}  показов {i:>6.0f}  кликов {c:>4.0f}  "
                      f"CTR {cr:>5.1f}%  {q}")

        # Высокий CTR — доказательство кликабельного интента.
        good = [q for q in qs if q[2] >= 20 and q[1] / q[2] >= 0.05]
        if good:
            good.sort(key=lambda x: -(x[1] / x[2]))
            print(f"\nВысокий CTR ({len(good)}) — такой интент работает, брать похожее:")
            for q, c, i, p in good[:8]:
                print(f"  CTR {c/i*100:>5.1f}%  поз {p:>4.1f}  кликов {c:>4.0f}  {q}")


def cmd_pages(_=None):
    blocks = [(l, k, d) for l, k, c, d in load_inbox() if "page" in c and "query" not in c]
    docs = repo_slugs()
    if not blocks:
        print("# выгрузок со страницами не нашёл")
        return
    seen = {}
    for label, kind, data in blocks:
        rows = []
        for d in data:
            u = str(d.get("page", "")).strip()
            if not u:
                continue
            rows.append((u, num(d.get("clicks")), num(d.get("impr")), num(d.get("pos"))))
        if not rows:
            continue
        print(f"\n=== {kind} — {label}")
        print(f"страниц {len(rows)}, кликов {sum(r[1] for r in rows):.0f}")
        rows.sort(key=lambda x: -x[1])
        print("\nЛидеры по кликам:")
        for u, c, i, p in rows[:8]:
            cr = c / i * 100 if i else 0
            print(f"  {c:>5.0f} кликов  CTR {cr:>5.1f}%  {slug_of(u) or u}")
        zero = [r for r in rows if r[1] == 0 and r[2] >= 50]
        if zero:
            zero.sort(key=lambda x: -x[2])
            print(f"\nПоказы без кликов ({len(zero)}) — сниппет или интент не работают:")
            for u, c, i, p in zero[:8]:
                print(f"  показов {i:>6.0f}  поз {p:>4.1f}  {slug_of(u) or u}")
        for u, c, i, p in rows:
            s = slug_of(u)
            if s:
                seen[s] = seen.get(s, 0) + c

    if docs and seen:
        ext = set()
        f = ROOT / "external-slugs.txt"
        if f.exists():
            for line in f.read_text(encoding="utf-8").splitlines():
                s = line.split("#")[0].strip()
                if s:
                    ext.add(s)
        print(f"\n=== Сверка с репозиторием ({len(docs)} статей)")
        if len(seen) < len(docs) * 0.5:
            print(f"В выгрузке всего {len(seen)} страниц против {len(docs)} в репозитории.")
            print("Похоже, экспорт неполный (обрезан лимитом строк) — вывод о "
                  "непроиндексированных страницах делать нельзя.")
            print("Перевыгрузи без лимита, тогда сверка заработает.")
        else:
            silent = sorted(s for s in docs if s not in seen)
            if silent:
                print(f"Нет ни одного показа ({len(silent)}) — вероятно, не в индексе:")
                print("  " + ", ".join(silent[:25]))
                if len(silent) > 25:
                    print(f"  …и ещё {len(silent)-25}")
            else:
                print("Все статьи репозитория показываются в выдаче.")
        ghost = sorted(s for s in seen if s and s not in docs and s not in ext)
        if ghost:
            print(f"Есть в выдаче, но нет ни в репозитории, ни в external-slugs.txt "
                  f"({len(ghost)}) — созданы вручную в WP, стоит добавить в список:")
            print("  " + ", ".join(ghost[:15]))


def cmd_digest(_=None):
    print(f"# Сводка по выгрузкам, {date.today().isoformat()}")
    print(f"# Источник: {INBOX.relative_to(ROOT)}. Данные сервисов — как выгружены.\n")
    cmd_files()
    cmd_queries()
    cmd_pages()
    print("\n# Что делать со сводкой:")
    print("#   «Топ-3 и ноль кликов» — такие темы больше не брать")
    print("#   «Близко к топу» — дописать разделы в существующую статью, не плодить спутники")
    print("#   «Высокий CTR» — образец интента для новых тем")
    print("#   «Нет ни одного показа» — проверить входящие ссылки: mdi.py in <слаг>")


def cmd_archive(_=None):
    files = [p for p in INBOX.iterdir() if p.is_file()] if INBOX.exists() else []
    if not files:
        sys.exit("Нечего архивировать.")
    dst = ARCHIVE / date.today().isoformat()
    dst.mkdir(parents=True, exist_ok=True)
    for p in files:
        shutil.move(str(p), str(dst / p.name))
    print(f"Перенесено {len(files)} файлов в {dst.relative_to(ROOT)}")


CMDS = {"files": cmd_files, "digest": cmd_digest, "queries": cmd_queries,
        "pages": cmd_pages, "archive": cmd_archive}

if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "digest"
    if cmd not in CMDS:
        print(__doc__)
        sys.exit(2)
    CMDS[cmd]()
